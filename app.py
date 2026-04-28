# -*- coding: utf-8 -*-
"""
freee人事労務API × Streamlit
勤怠データを取得してExcelに転記するWEBアプリ
"""

import os
import requests
import calendar
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date, time, timedelta
from urllib.parse import urlencode
from dotenv import load_dotenv
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import streamlit as st
from openpyxl import load_workbook

import ui

# 設定読込: ローカルは .env、Streamlit Cloud は st.secrets を使用
load_dotenv()


def _cfg(key: str) -> str | None:
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.getenv(key)


CLIENT_ID = _cfg("CLIENT_ID")
CLIENT_SECRET = _cfg("CLIENT_SECRET")
REDIRECT_URI = _cfg("REDIRECT_URI")

TEMPLATE_FILE = "YYYY年_MM月_作業実施報告書_SMHC_苗字 名前.xlsx"


# =========================
# 共有リソース（プロセス内で1回だけ生成）
# =========================

@st.cache_resource
def _http_session() -> requests.Session:
    """freee API 向け共有セッション。429/5xx で自動リトライ。"""
    s = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    s.mount("https://", adapter)
    return s


@st.cache_resource
def _template_bytes() -> bytes:
    """テンプレートExcelを起動時に1度だけメモリへ読み込む。
    以降はディスクI/Oを行わないので、ローカルでテンプレートを開いていても影響なし。"""
    with open(TEMPLATE_FILE, "rb") as f:
        return f.read()


# =========================
# OAuth関連処理
# =========================

def get_auth_url():
    params = urlencode({
        "client_id": CLIENT_ID,
        "response_type": "code",
        "redirect_uri": REDIRECT_URI,
        "scope": "hr",
    })
    return f"https://accounts.secure.freee.co.jp/public_api/authorize?{params}"


def save_token(token_data):
    st.session_state["token"] = token_data


def load_token():
    return st.session_state.get("token")


def get_token_from_code(code):
    url = "https://accounts.secure.freee.co.jp/public_api/token"
    data = {
        "grant_type": "authorization_code",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code": code,
        "redirect_uri": REDIRECT_URI,
    }
    res = _http_session().post(url, data=data)
    token = res.json()
    token["expires_at"] = (datetime.now() + timedelta(seconds=token["expires_in"])).timestamp()
    save_token(token)
    return token


def refresh_token(token):
    url = "https://accounts.secure.freee.co.jp/public_api/token"
    data = {
        "grant_type": "refresh_token",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "refresh_token": token["refresh_token"],
    }
    res = _http_session().post(url, data=data)
    new_token = res.json()
    new_token["expires_at"] = (datetime.now() + timedelta(seconds=new_token["expires_in"])).timestamp()
    save_token(new_token)
    return new_token


def get_valid_token():
    token = load_token()
    if not token:
        return None
    if datetime.now().timestamp() > token["expires_at"]:
        token = refresh_token(token)
    return token["access_token"]


# =========================
# ユーザー情報取得
# =========================

def get_user_info(access_token):
    url = "https://api.freee.co.jp/hr/api/v1/users/me"
    headers = {"Authorization": f"Bearer {access_token}"}
    res = _http_session().get(url, headers=headers)
    if res.status_code != 200:
        st.error(f"ユーザー情報取得エラー: {res.text}")
        return None
    return res.json()


def split_display_name(display_name):
    """display_name を空白区切りで苗字・名前に分割。
    空白がなければ全体を苗字として返す。"""
    if not display_name:
        return None, None
    parts = display_name.strip().split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return parts[0], None


# =========================
# API処理
# =========================

_HHMM_RE = __import__("re").compile(r"(\d{1,2}):(\d{2})")


def _parse_hhmm(s):
    """freeeの様々な時刻表現から time オブジェクトを返す
    対応例:
      '2026-04-01T09:00:00.000+09:00' (ISO8601 / T区切り / タイムゾーン付)
      '2026-04-01 09:00:00'           (空白区切り)
      '09:00:00'                      (時刻のみ)
    """
    if not s:
        return None
    s = str(s).strip()
    # 日付と時刻を分離 (T または 空白)
    if "T" in s:
        time_part = s.split("T", 1)[1]
    elif " " in s:
        time_part = s.split(" ", 1)[1]
    else:
        time_part = s
    m = _HHMM_RE.match(time_part)
    if not m:
        return None
    try:
        return time(int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def _break_total_hours(break_records):
    """break_records 配列から合計休憩時間(時間, 小数)を計算"""
    total_minutes = 0
    for b in break_records or []:
        t1 = _parse_hhmm(b.get("clock_in_at"))
        t2 = _parse_hhmm(b.get("clock_out_at"))
        if t1 and t2:
            total_minutes += max(0, (t2.hour * 60 + t2.minute) - (t1.hour * 60 + t1.minute))
    return round(total_minutes / 60, 2) if total_minutes else 0


def get_work_records(year, month, access_token, company_id, employee_id):
    """各日ごとに勤怠データを取得 (GET /employees/{id}/work_records/{date})

    並列リクエストで高速化。同時利用者数 × max_workers ≦ freee 側のレート制限を
    超えないよう、max_workers は 4 に抑制（リトライ機構と併用）。"""
    headers = {"Authorization": f"Bearer {access_token}"}
    params = {"company_id": company_id}
    days = calendar.monthrange(year, month)[1]
    dates = [f"{year}-{month:02d}-{d + 1:02d}" for d in range(days)]
    session = _http_session()

    def _fetch(date_str):
        url = (
            "https://api.freee.co.jp/hr/api/v1"
            f"/employees/{employee_id}/work_records/{date_str}"
        )
        return date_str, session.get(url, headers=headers, params=params)

    records = []
    errors = []
    progress = st.progress(0.0, text="勤怠データ取得中...")
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = [ex.submit(_fetch, d) for d in dates]
        for i, fut in enumerate(as_completed(futures)):
            date_str, res = fut.result()
            if res.status_code == 200:
                rec = res.json()
                if not rec.get("date"):
                    rec["date"] = date_str
                records.append(rec)
            elif res.status_code != 404:
                errors.append(f"{date_str}: status={res.status_code} {res.text[:120]}")
            progress.progress((i + 1) / days, text=f"{i + 1}/{days} 件取得")
    progress.empty()

    # 並列取得は順序が不定なので日付順に整列
    records.sort(key=lambda r: r.get("date", ""))

    if errors:
        st.warning("一部の日付でエラーが発生しました:\n" + "\n".join(errors[:5]))

    # デバッグ用: 有給/半休レコードを抽出して表示
    if records:
        leave_records = [r for r in records if _is_leave_record(r)]
        with st.expander(
            f"🔍 有給/半休レコード ({len(leave_records)} 件)",
            expanded=bool(leave_records),
        ):
            if leave_records:
                for r in leave_records:
                    st.write(f"**{r.get('date')}** - 判定: {detect_leave_type(r) or '(該当なし)'}")
                    st.json(r)
            else:
                st.caption("有給/半休に該当するレコードなし")

    return records


def _is_leave_record(r):
    return (
        (r.get("paid_holiday") or 0) != 0
        or (r.get("half_paid_holiday_mins") or 0) != 0
        or (r.get("half_special_holiday_mins") or 0) != 0
        or (r.get("hourly_paid_holiday_mins") or 0) != 0
        or (r.get("special_holiday") or 0) != 0
        or bool(r.get("half_holiday_type"))
        or bool(r.get("is_absence"))
    )


def detect_leave_type(r):
    """freee レコードから有給/午前休/午後休を判定
    Returns: "paid_full" | "am_half" | "pm_half" | None
    """
    paid = float(r.get("paid_holiday") or 0)
    half_type = (r.get("half_holiday_type") or "").lower()
    half_paid_mins = r.get("half_paid_holiday_mins") or 0

    # 全日有給
    if paid >= 1.0:
        return "paid_full"

    # 半休判定 (paid_holiday=0.5 または half_paid_holiday_mins>0)
    is_half = paid >= 0.5 or half_paid_mins > 0 or bool(half_type)
    if is_half:
        if "am" in half_type:
            return "am_half"
        if "pm" in half_type:
            return "pm_half"

    return None


def format_work_data(records):
    data = {}
    for r in records:
        d = r.get("date")
        if not d:
            continue
        data[d] = {
            "clock_in": _parse_hhmm(r.get("clock_in_at")),
            "clock_out": _parse_hhmm(r.get("clock_out_at")),
            "break": _break_total_hours(r.get("break_records")),
            "leave_type": detect_leave_type(r),
        }
    return data


LEAVE_REMARK = {
    "paid_full": "有給取得",
    "am_half": "午前休取得",
    "pm_half": "午後休取得",
}


def write_to_excel(year, month, data, employee_name, last_name=None, first_name=None, project_name=None):
    wb = load_workbook(BytesIO(_template_bytes()))
    ws = wb.active
    ws["K5"] = last_name
    ws["L5"] = first_name
    ws["K8"] = date(year, month, 1)  # date型で書き込み → L列の weekday書式(ddd)が機能する
    if project_name:
        ws["K3"] = project_name

    days = calendar.monthrange(year, month)[1]
    for i in range(days):
        row = 8 + i
        d = date(year, month, i + 1)
        date_str = d.strftime("%Y-%m-%d")
        record = data.get(date_str, {})
        clock_in = record.get("clock_in")
        clock_out = record.get("clock_out")
        break_hours = record.get("break")
        leave_type = record.get("leave_type")

        if leave_type == "paid_full":
            # 全日有給: 備考のみ記載、時刻・休憩は空欄
            ws[f"H{row}"] = LEAVE_REMARK[leave_type]
            ws[f"M{row}"] = None
            ws[f"N{row}"] = None
            ws[f"O{row}"] = None
        elif leave_type in ("am_half", "pm_half"):
            # 半休: 備考 + 始業/終業時刻 + 休憩(無ければ0.0)
            ws[f"H{row}"] = LEAVE_REMARK[leave_type]
            ws[f"M{row}"] = clock_in
            ws[f"N{row}"] = clock_out
            ws[f"O{row}"] = break_hours if break_hours else 0.0
        else:
            # 通常日
            ws[f"M{row}"] = clock_in
            ws[f"N{row}"] = clock_out
            # 出退勤とも無し → 休憩時間欄を空欄 (0.0 を表示しない)
            if not clock_in and not clock_out:
                ws[f"O{row}"] = None
            else:
                ws[f"O{row}"] = break_hours

    safe_name = (employee_name or "未設定").strip()
    file_name = f"{year}_{month}月_作業実施報告書_SMHC_{safe_name}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return file_name, buffer


# =========================
# Streamlit UI
# =========================

st.set_page_config(
    page_title="freee 勤怠レポート",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# 現在時刻（時計アニメ初期位置に使用）
now = datetime.now()
year = now.year
month = now.month

# スタイル読込（styles.css + 時計ディレイの CSS 変数）
st.markdown(ui.load_styles(now), unsafe_allow_html=True)

# 背景ブロブ + ヒーロー
st.markdown(ui.ambient_blobs(), unsafe_allow_html=True)
st.markdown(ui.hero(), unsafe_allow_html=True)

# トークン
access_token = get_valid_token()

# OAuth リダイレクトで戻ってきた場合、URLクエリ ?code=... を自動で拾って認証する
if not access_token:
    qp_code = st.query_params.get("code")
    if qp_code:
        try:
            with st.spinner("認証中..."):
                get_token_from_code(qp_code)
            st.query_params.clear()  # URLから認可コードを消す
            st.success("認証完了")
            st.rerun()
        except Exception as e:
            st.error(f"認証エラー: {e}")
            st.query_params.clear()

# サイドバー
with st.sidebar:
    st.markdown("### Status")
    st.markdown(ui.status_pill(access_token is not None), unsafe_allow_html=True)

    st.markdown("### Period")
    st.markdown(ui.icon_row_period(year, month), unsafe_allow_html=True)
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)
    st.caption(f"自動判定: {now.strftime('%Y-%m-%d')}")

    st.markdown("### Security")
    st.markdown(ui.icon_row_security(), unsafe_allow_html=True)

    st.markdown("---")

    if access_token:
        if st.button("ログアウト", use_container_width=True):
            st.session_state.pop("token", None)
            st.rerun()

    st.markdown(ui.footer_note(), unsafe_allow_html=True)


# メイン
if not access_token:
    st.markdown(ui.stepper(active_step=1), unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2.2, 1])
    with col_c:
        with st.container(border=True):
            st.markdown(ui.shield_scene(), unsafe_allow_html=True)
            st.markdown(ui.auth_card_header(), unsafe_allow_html=True)

            auth_url = get_auth_url()

            st.markdown(
                """
<style>
.auth-instr-line {
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: clip !important;
    font-size: 0.68rem !important;
    line-height: 1.6 !important;
    margin: 0.4rem 0 !important;
    letter-spacing: -0.01em !important;
    display: block !important;
    width: 100% !important;
}
.auth-instr-line strong { font-size: inherit !important; white-space: nowrap !important; }
</style>
<p class="auth-instr-line"><strong>「freee認証ページを開く」</strong>をクリック → freeeで許可すると自動で認証されます。</p>
""",
                unsafe_allow_html=True,
            )
            st.link_button("freee認証ページを開く", auth_url, use_container_width=True)

            with st.expander("自動で認証されない場合（手動入力）"):
                st.caption("リダイレクト先URL `?code=xxxxx` の xxxxx 部分を貼り付けてください。")
                code = st.text_input(
                    "認可コード",
                    placeholder="ここに認可コードを入力",
                    label_visibility="collapsed",
                )

            if st.button("認証する", type="primary", use_container_width=True):
                if not code:
                    st.warning("認可コードを入力してください")
                else:
                    with st.spinner("認証中..."):
                        get_token_from_code(code)
                    st.success("認証完了")
                    st.rerun()

else:
    user_info = get_user_info(access_token)
    if not user_info:
        st.stop()

    companies = user_info.get("companies", [])
    if not companies:
        st.error("所属会社が見つかりません")
        st.stop()

    # freee HR API の仕様:
    #   companies[].name         = 会社名
    #   companies[].display_name = ユーザー本人の従業員表示名（氏名）
    options = {
        (c.get("name") or f"会社ID:{c['id']}"): c
        for c in companies
    }

    m1, m2, m3 = st.columns(3)
    user_label = user_info.get("email") or f"User #{user_info.get('id', '-')}"
    m1.metric("User", user_label)
    m2.metric("対象月", f"{year}年{month:02d}月")
    m3.metric("所属会社", f"{len(companies)} 社")

    with st.container(border=True):
        st.markdown(ui.report_card_header(), unsafe_allow_html=True)

        label = st.selectbox("対象会社", list(options.keys()))
        selected = options[label]
        company_id = selected["id"]
        employee_id = selected.get("employee_id")

        project_name = st.text_input("契約名 (任意)")

        if not employee_id:
            st.error("この会社には従業員IDが紐付いていません")
        else:
            st.markdown(ui.info_row_ids(company_id, employee_id), unsafe_allow_html=True)

            if st.button("データ取得 & Excel作成", type="primary", use_container_width=True):
                with st.status("レポートを生成しています...", expanded=True) as status:
                    st.write("① 従業員情報を取得中...")
                    # /users/me の companies[].display_name から氏名を取得し分割
                    employee_name = selected.get("display_name")
                    last_name, first_name = split_display_name(employee_name)
                    st.write(f"　→ {employee_name or '(取得失敗)'} (苗字: {last_name or '-'} / 名前: {first_name or '-'})")

                    st.write("② 勤怠データを取得中...")
                    records = get_work_records(year, month, access_token, company_id, employee_id)
                    st.write(f"　→ {len(records)} 件のレコードを取得")

                    st.write("③ データを整形中...")
                    data = format_work_data(records)

                    st.write("④ Excelファイルを生成中...")
                    file_name, file_buffer = write_to_excel(
                        year, month, data, employee_name,
                        last_name=last_name, first_name=first_name,
                        project_name=project_name,
                    )
                    st.write(f"　→ {file_name}")

                    status.update(label="レポート生成完了", state="complete", expanded=False)

                st.download_button(
                    label="Excelをダウンロード",
                    data=file_buffer,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

st.markdown(ui.footer_note("© <b>freee Kintai Report</b> · Streamlit で構築"), unsafe_allow_html=True)
